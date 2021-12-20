import random
import string
import openpyxl as op
import os, sys
import math

questionnaireCount = 121
###################产生2个以上随机整数###################
###################第一个数随机产生，第二个使用平均数求出###################
#count 数字的个数
#average 平均数
#begin 起始区间
#end 结束区间

def int_random_old (average, begin, end):

    numarr = [];

    while (1):

        num_first = random.randint(begin, end);
        num_second = average * 2 - num_first;

        if (num_second >= begin and num_second <= end):
            numarr.append(num_first);
            numarr.append(num_second);
            break
        # else:
            # print("num_second: " + str(num_second))

    return numarr;


def int_random (average, begin, end):

    #print "wzh_random"
    # numarr = [0 for x in range(2)];
    numarr = [];
    i = 0;
    # while (1):
 
    #   num_first = random.randrange(begin, end);
    num_first = random.randint(begin, end);

    #第二个数
    num_second = average * 2 - num_first;

    numarr.append(num_first);
  
   
    if (num_second >= begin and num_second <= end):  
        numarr.append(num_second);
    else:
        # 继续分配,分为三个数
        # 12 -1 ==11 ; 11+4 = 15
        # 15 -3 ==12 ; 12+5 = 17
        rest_val_temp = average * 3 - num_first

        restTwo = re_random(rest_val_temp, begin, end,average)
        numarr.extend(restTwo)

    return numarr;

def re_random (rest_val, begin, end,average):
    numarr = [];
    random_number = random.uniform(0,1)
    if rest_val <= 10:
        if random_number <= 0.5:
            # 大值与小值
            numarr.append(end)
            numarr.append(rest_val - end)
        else:
            # 平均分
            if (rest_val % 2) == 0:
                #偶数
                numarr.append(rest_val/2)
            else:
                #奇数
                numarr.append(rest_val//2)
                numarr.append(rest_val//2 + 1)
    else:
        rest_val2 = rest_val + average
        mean_temp =    rest_val2//3
        numarr.append(mean_temp)
        numarr.append(mean_temp)
        numarr.append(rest_val2-mean_temp*2)

    """     
    for i in range(len(numarr)):
        if numarr[i] > end: 
    """

            

    return numarr;





###################产生随机数###################
###################第一个数随机产生，第二个使用平均数求出###################
#count 数字的个数
#average 平均数
#begin 起始区间
#end 结束区间
def float_random (count, average, begin, end):

    #print "wzh_random"
    numarr = [0 for x in range(2)];
    i = 0;
    while (1):
 
      num = random.uniform(begin, end);
      #取两位小数
      num_first = round(num, 2);

      #第二个数
      num_second = average * 2 - num_first;

      if (num_second >= begin and num_second <= end):
          numarr[i] = num_first;
          i = i + 1;
          numarr[i] = num_second;
          break

    return numarr;



###################写文件###################
def write_file (filename, content):
    fo = open (filename, "rb+");
    fo.write(bytes(content,'utf-8'));
    # fo.write(content);
    fo.close();

def show_list (list):
    for i in list:
        print(i, end=' ')
    print();
    
    t_sum = 0   
    t_average = 0.0

    t_sum = sum(list)
    t_average = round(t_sum/questionnaireCount,4)

    print("列表的和：" + str(t_sum))
    print("数量："+ str(len(list)));
    print("列表的平均数：" + str(t_average))

###################主函数调用产生整形随机数###################

def dataModify(list):
    # 生成一个1到4到随机数


    lastIndex = len(list) - 1
    num_begin = random.randint(1, 4);

    """ 
    num_list = [] 
    num_list.append(num_begin)
    num_list.append(num_begin+1)
    num_list.append(num_begin+2) """

    swapPositions(list,num_begin,lastIndex-num_begin)
    swapPositions(list,num_begin+1,lastIndex-num_begin-1)
    swapPositions(list,num_begin+2,lastIndex-num_begin-2)

    return list
def swapPositions(list, pos1, pos2):
     
    list[pos1], list[pos2] = list[pos2], list[pos1]
    return list

def test_random_int(ava):
    
    realAverage = ava

    begin = math.floor(ava)
    end = math.ceil(ava)

    """
    if(int(ava) ==4):
        begin = 4
        end = 5
    elif(int(ava) == 3):
        begin = 3
        end = 4
    elif(int(ava) == 2):
        begin = 2
        end = 3
      
    #向上取整
    print math.ceil(num)

    #向下取整
    print math.floor(num)

    #简单直接取整
    print int(num)

    #四舍五入取整
    print round(num) 
    """
    
    # 对平均数取整以后，要对平均数进行调整 5的话不出现begin=2, 3的话end =4
    average = round(realAverage)
    print(str(realAverage) + " >>>>>> " + str(average) )
    count = questionnaireCount;
    # begin = 3;
    # end = 4;
    numarr_count = 0;
    numarr = [0 for x in range(count)];

    
    print("begin："+ str(begin))
    print("end："+ str(end))

    # for i in range (count // 2):
    while(numarr_count < count):
        templist = int_random_old (average, begin, end)
        j = 0;
        for j in range (len(templist)):
            # numarr_count 有可能大于120
            if numarr_count < count:
                numarr[numarr_count] = templist[j]
                numarr_count += 1
            else:
                print('注意!!!!!: numarr_count: ' + str(numarr_count))
                break;

    
    # print("修正前 ######：")
    # show_list (numarr)

    # 数据修正开始
    #if (count % 2) != 0:
     #   numarr[count-1] = random.randint(begin, end);

    t_average2 = round(sum(numarr)/count,4)
    # 平均数差值
    gapValue = round( realAverage - t_average2,4)
    if(  gapValue != 0):
        # 差了多少没有分配
        undistributed = round(count * gapValue,4)
        print("undistributed: " + str(undistributed))
        # 四舍五入取整
        undistributed = round(undistributed)

        # 分配到每个数里去,多退少补
        if(undistributed > 0):
            for x in range(len(numarr)):
                # undistributed 用完的时候跳出
                if(undistributed <= 0):
                    print("undistributed 剩余: " + str(undistributed))
                    break;
                if (numarr[x] < end):
                    numarr[x]  +=1
                    undistributed -=1
        else:
             for x in range(len(numarr)):
                # undistributed 用完的时候跳出
                if(undistributed >= 0):
                    print("undistributed 剩余: " + str(undistributed))
                    break;
                if (numarr[x] > begin or numarr[x] == 5):
                    numarr[x]  -=1
                    undistributed +=1
    # 从小到大排序
    numarr = sorted(numarr)
    # 数据打乱
    # random.shuffle(numarr);

    # 排序以后在spss里分析，信度会过高，效度分析时关联性太强，无法显示KMO值，因此要稍微弄乱一点
    # 随机在前五 和 倒数前五个数中 分别抽取3个， 相互调换位置
    numarr = dataModify(numarr)
    return numarr;

          

    # if(undistributed <= end):
    # numarr[count-1] = int(undistributed)
    # else:


             
    content = '';
    #打乱排序
    # print("修正后 ######：")
    # print("数据未打乱：");
    # show_list (numarr)
    """     
    random.shuffle(numarr);
        print("数据打乱：");
        show_list (numarr) 
    """
    for i in numarr:
        content = content + "\n" + str(i)
    #print content;
    #追加写入文件
    filename = "test3.txt";
    # print("文件名称：",filename);
    write_file (filename, content)
    write_file (filename, "\n");

###################主函数调用产生实型随机数###################
#40个数字，平均数400，363 - 429 之间
def test_random_float():
    count = 40;
    average = 402;
    begin = 363;
    end = 429;
    numarr_count = 0;
    numarr = [0 for x in range(count)];
    for i in range (count // 2):
        list = float_random (40, 402, 363, 429)
        j = 0;
        for j in range (len(list)):
             numarr[numarr_count] = list[j];
             numarr_count += 1;
    content = '';
    #打乱排序
    print("数据未打乱：");
    show_list (numarr)
    random.shuffle(numarr);
    print("数据打乱：");
    show_list (numarr)
    for i in numarr:
        content = content + ' ' + str(i);
    #print content;
    #追加写入文件
    filename = "test2.txt";
    print("文件名称：",filename);
    write_file (filename, content)
    write_file (filename, "\n");

#调用测试产生整形随机数
# test_random_int(3.6158);
#调用测试产生实型随机数
# test_random_float();
# sys.exit(0)
#expectList = [4.7678,3.6186,4.2126,4.645,3.6186,4.2086,4.3846,4.2955,3.9666,4.0586,4.2747,4.5946,4.2701,3.9326,4.2927,4.4814,3.9269,4.5673,4.2785,3.8283,4.4761,4.4532,4.4108]

perciveList = [2.5332,4.258,3.437,2.6852,4.1787,3.241,3.7591,3.6448,3.372,3.63,4.258,2.8037,3.9104,3.6703,3.8763,3.997,3.7745,3.1142,3.9463,3.6831,3.1383,3.2094,3.3453]

bg = op.load_workbook(r"data_1.xlsx")    

for colIndex in range(len(perciveList)):

    resultColum = test_random_int(perciveList[colIndex]) 
    
    sheet = bg["perciveList"]                             		 
    for row in range(len(resultColum)):						
        sheet.cell(row+1 , colIndex + 1, resultColum[row])	# sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save("data_1.xlsx")            			# 对文件进行保存         


# https://blog.csdn.net/l734971107/article/details/109635668